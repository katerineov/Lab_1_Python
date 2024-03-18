# импортируем необходимые библиотеки
import openpyxl
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

data = {
        'device': 'OS X Chrome v.88.0.4389.90 3a7a0',
        'app_version': '870'
    }

headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
    }

# записываем "базовую" ссылку, через которую мы будем листать страницы с объявлениями
base_url = 'https://auto.drom.ru/all/page{}/'

def parsedrom():
    page_number = 8 # номер страницы, с которой начнется парсинг
    car_name = [] #  список, который будет хранить названия автомобилей, извлеченных с сайта
    car_links = [] # список, который будет хранить ссылки на страницы с объявлениями
    while len(car_name) < 20: # будет продолжаться до тех пор, пока в списке не будет не менее 20 элементов.
        # передаем необходимый URL адрес
        url = base_url.format(page_number)
        # используем библиотеку requests для отправки GET-запроса по указанному URL-адресу и сохраняем ответ в переменной response
        response = requests.get(url)

        if response.status_code != 200:
            print(f'Error accessing page {page_number}') # если код состояния не равен 200, выводит сообщение об ошибке, указывая текущий номер страницы
            break

        soup = BeautifulSoup(response.text, "html.parser") # передаем страницу в bs4


        names = soup.findAll('div', class_='css-16kqa8y e3f4v4l2') # находим  контейнер с нужным классом

        count = 0
        for name in names:
            if count == 20:
                break
            # проверяем, содержит ли текущий элемент элемент div с классом 'css-1ajc6qi e3f4v4l1'. если содержит, то увеличиваем count и переходим к следующей итерации
            if name.find('div', class_='css-1ajc6qi e3f4v4l1'):  # проверка - закреплено ли объявление (по титл)
                count += 1
                continue
            else:
                car_name.append(name.text)

        # ищем все ссылки (<a>) с классом 'css-4zflqt e1huvdhj1' на странице и записываем их в переменную link
        link = soup.findAll('a', class_='css-4zflqt e1huvdhj1')
        for links in link:
            # если количество ссылок равно количеству добавленных имен автомобилей, прерываем цикл
            if len(car_links) == len(car_name):
                break
            car_link = links.get('href') # получаем адрес ссылки (href) и записываем его в переменную car_link
            car_links.append(car_link)
        car_links = car_links[count:] # jбрезаем список car_links, начиная с позиции count

        page_number += 1

    # cоздаем файл и записываем в него список списков с именами автомобилей и соответствующими им ссылками
    file([[_ for _ in car_name], [_ for _ in car_links]])

def file(info):
    excel_file = load_workbook('dromads.xlsx') # загружаем файл excel dromads.xlsx в переменную excel_file
    excel_file_page1 = excel_file['all_ads'] # получаем доступ к листу 'all_ads'
    excel_file_page1.delete_rows(2, excel_file_page1.max_row) # удаляем все строки, начиная со второй и до конца листа
    excel_file.save('dromads.xlsx') # сохраняем изменения
    for i in range(len(info[0])): # запускаем цикл по длине списка info[0].
        excel_file_page1.append(([info[0][i], info[1][i]])) # добавляем 2 элемента из списка info (один из info[0][i] и второй из info[1][i]) в конец листа excel
    excel_file.save('dromads.xlsx') # сохраняем изменения
    excel_file.close() # обязательно закрываем файл
